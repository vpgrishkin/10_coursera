from random import sample

import requests
from lxml import etree
from openpyxl import Workbook
from bs4 import BeautifulSoup


XML_COURSES_URL = 'https://www.coursera.org/sitemap~www~courses.xml'
DEFAULT_EXCEL_FILE_NAME = 'courses.xlsx'
DEFAULT_COURSES_COUNT = 20


def get_courses_list(url=XML_COURSES_URL):
    xml = requests.get(url).content
    root = etree.XML(xml)
    return [link.text for link in root.iter('{*}loc')]


def get_course_info(course_slug):
    soup = BeautifulSoup(course_slug, 'html.parser')
    title = soup.find('h1', class_='title').text
    start_date = soup.find('div', class_='startdate').text if soup.find('div', class_='startdate') else None
    start_date = start_date.split(maxsplit=1)[1] if start_date else None
    languages = soup.find('div', class_='language-info').text
    language = languages.split(',')[0]
    duration_in_weeks = len(soup.find_all('div', class_='week'))
    rating_tag = soup.find('div', class_='ratings-text')
    if rating_tag and rating_tag.text:
        rating = rating_tag.text.split()[0]
    else:
        rating = None
    return {'title': title,
            'starting_date': start_date,
            'language': language,
            'duration_in_weeks': duration_in_weeks,
            'rating': rating}


def output_courses_info_to_xlsx(filepath, courses_info):
    excel_workbook = Workbook()
    sheet = excel_workbook.active
    sheet.title = 'Coursera'
    sheet['A1'] = 'Course title'
    sheet['B1'] = 'Starting date'
    sheet['C1'] = 'Language'
    sheet['D1'] = 'Duration (weeks)'
    sheet['E1'] = 'Rating'
    for row, course in enumerate(courses_info, 2):
        sheet.cell(row=row, column=1, value=course['title'])
        sheet.cell(row=row, column=2, value=course['starting_date'])
        sheet.cell(row=row, column=3, value=course['language'])
        sheet.cell(row=row, column=4, value=course['duration_in_weeks'])
        if course['rating']:
            sheet.cell(row=row, column=5, value=course['rating'])
        else:
            sheet.cell(row=row, column=5, value='No rating')
    excel_workbook.save(filepath)


if __name__ == '__main__':
    print('Start loading courses from {}'.format(XML_COURSES_URL))
    all_courses_urls_list = get_courses_list()
    random_courses_urls = sample(all_courses_urls_list, DEFAULT_COURSES_COUNT)
    print('Get random courses urls list: \n {}'.format(random_courses_urls))

    courses_raw_pages = [requests.get(course_url).content for course_url in random_courses_urls]
    courses_info = [get_course_info(course_raw_page) for course_raw_page in courses_raw_pages]

    output_courses_info_to_xlsx(DEFAULT_EXCEL_FILE_NAME, courses_info)
    print('Start saving courses to {}'.format(DEFAULT_EXCEL_FILE_NAME))
    print('Done')
